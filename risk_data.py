import openpyxl
from datetime import datetime

def check_risk_data(filename="risk_data.xlsx"):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print(f"Error: The file '{filename}' was not found.")
        return [], []
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Error: The file '{filename}' is not a valid Excel file.")
        return [], []

    ws = wb.active
    
    risks = []
    errors = []

    expected_headers = [
        "Risk ID", "Description", "Initial Probability", "Occurrence Time (Day)", 
        "Cost PERT (min)", "Cost PERT (most likely)", "Cost PERT (max)",
        "Delay PERT (min)", "Delay PERT (most likely)", "Delay PERT (max)",
        "Dependent Risks", "Probability Adjustment"
    ]

    actual_headers = [cell.value for cell in ws[1]]
    missing_columns = set(expected_headers) - set(actual_headers)
    if missing_columns:
        for column in missing_columns:
            errors.append(f"Error: Column '{column}' is missing from the Excel file.")
        print("Errors found in the Excel structure. Please correct these before proceeding with data validation.")
        return errors, []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < len(expected_headers):
            errors.append(f"Error: Row {ws.index(row)} has fewer columns than expected.")
            continue
        
        try:
            risk = {
                "Risk ID": row[0],
                "Description": row[1],
                "Initial Probability": float(row[2]),
                "Occurrence Time": int(row[3]),
                "Cost PERT": [
                    float(str(row[4]).replace(',', '.')),
                    float(str(row[5]).replace(',', '.')),
                    float(str(row[6]).replace(',', '.'))
                ],
                "Delay PERT": [int(row[7]), int(row[8]), int(row[9])],
                "Dependent Risks": row[10].split(", ") if row[10] else [],
                "Probability Adjustment": float(row[11]) if row[11] else 0
            }
            risks.append(risk)
        except ValueError as e:
            errors.append(f"Error processing row {ws.index(row)}: {e}")

    risks.sort(key=lambda x: x["Occurrence Time"])

    for i, risk in enumerate(risks):
        if risk["Dependent Risks"]:
            dep_risks = risk["Dependent Risks"]
            for dep_risk in dep_risks:
                dep_index = next((i for i, r in enumerate(risks) if r["Risk ID"] == dep_risk), None)
                if dep_index is None:
                    errors.append(f"Error: {risk['Risk ID']} has non-existent dependent risk {dep_risk}.")
                elif dep_index <= i:
                    errors.append(f"Error: {risk['Risk ID']} has a dependency {dep_risk} that doesn't occur in the future.")

        if risk["Dependent Risks"] and not risk["Probability Adjustment"]:
            errors.append(f"Error: {risk['Risk ID']} has dependent risks but no probability adjustment.")
        elif not risk["Dependent Risks"] and risk["Probability Adjustment"]:
            errors.append(f"Error: {risk['Risk ID']} has a probability adjustment but no dependent risks.")
        elif risk["Probability Adjustment"]:
            try:
                prob_adj = float(risk["Probability Adjustment"])
                if not (-0.20 <= prob_adj <= 0.20 and abs(prob_adj) >= 0.05):
                    errors.append(f"Error: {risk['Risk ID']} has a probability adjustment {prob_adj} outside the valid range [-0.20, -0.05] or [0.05, 0.20].")
                
                is_opportunity = "Opportunity" in risk["Description"]
                if (is_opportunity and prob_adj > 0) or (not is_opportunity and prob_adj < 0):
                    errors.append(f"Error: {risk['Risk ID']} has an incorrect probability adjustment sign for its type (opportunity/threat).")
            except ValueError:
                errors.append(f"Error: {risk['Risk ID']} has an invalid probability adjustment value.")

    if errors:
        print("The following errors were found:")
        for error in errors:
            print(error)
    else:
        print("No errors found. The risk data is consistent.")

    with open("risk_check_log.txt", "a") as log_file:
        log_file.write(f"Check performed on {datetime.now()}\n")
        if errors:
            log_file.write("Errors found:\n")
            for error in errors:
                log_file.write(f"{error}\n")
        else:
            log_file.write("No errors found.\n")
        log_file.write("\n")

    return errors, risks

def read_and_validate_data(filename="risk_data.xlsx"):
    errors, risks = check_risk_data(filename)
    if errors:
        print("Errors found in the data. Please correct them before proceeding.")
        for error in errors:
            print(error)
        return None
    return risks

if __name__ == "__main__":
    check_risk_data()