import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

def save_matrices_to_excel(cumulative_cost_matrix, cumulative_delay_matrix, cost_bin_size, delay_bin_size, min_cost, min_delay, output_file):
    # Create a Pandas Excel writer using XlsxWriter as the engine
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        wb = writer.book

        # Create cost labels
        cost_labels = [min_cost + i * cost_bin_size for i in range(cumulative_cost_matrix.shape[1])]
        
        # Convert cumulative cost matrix to DataFrame
        cost_df = pd.DataFrame(cumulative_cost_matrix, columns=cost_labels)
        cost_df.index.name = 'Project Day'
        cost_df.columns.name = 'Cumulative Cost'

        # Write cumulative cost matrix to Excel
        cost_df.to_excel(writer, sheet_name='Cumulative Cost Matrix')
        
        # Add explanation
        ws = wb['Cumulative Cost Matrix']
        ws['A1'] = "Cumulative Cost Matrix"
        ws['A1'].font = Font(bold=True)
        ws['A2'] = "Each cell represents the number of iterations where the cumulative cost reached the column value by the row day."

        # Create delay labels
        delay_labels = [min_delay + i * delay_bin_size for i in range(cumulative_delay_matrix.shape[1])]
        
        # Convert cumulative delay matrix to DataFrame
        delay_df = pd.DataFrame(cumulative_delay_matrix, columns=delay_labels)
        delay_df.index.name = 'Project Day'
        delay_df.columns.name = 'Cumulative Delay'

        # Write cumulative delay matrix to Excel
        delay_df.to_excel(writer, sheet_name='Cumulative Delay Matrix')
        
        # Add explanation
        ws = wb['Cumulative Delay Matrix']
        ws['A1'] = "Cumulative Delay Matrix"
        ws['A1'].font = Font(bold=True)
        ws['A2'] = "Each cell represents the number of iterations where the cumulative delay reached the column value by the row day."

    print(f"Cumulative matrices saved to '{output_file}'.")

    # Print parameters for clarity
    print(f"Cumulative Cost Matrix Parameters:")
    print(f"  Min Cost: {min_cost}")
    print(f"  Cost Bin Size: {cost_bin_size}")
    print(f"  Number of Cost Bins: {cumulative_cost_matrix.shape[1]}")

    print(f"Cumulative Delay Matrix Parameters:")
    print(f"  Min Delay: {min_delay}")
    print(f"  Delay Bin Size: {delay_bin_size}")
    print(f"  Number of Delay Bins: {cumulative_delay_matrix.shape[1]}")