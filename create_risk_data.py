import openpyxl
from openpyxl.styles import Font
import random
import math

def generate_risk_data(num_risks=15, project_duration_days=365):
    risks = []
    num_opportunities = max(1, int(num_risks * 0.2))  # Ensure at least 1 opportunity, but no more than 20%
    
    for i in range(1, num_risks + 1):
        is_opportunity = i <= num_opportunities  # First 20% (rounded down) are opportunities
        
        def generate_pert_values(is_opportunity, min_val, max_val, is_cost=True):
            if is_opportunity:
                max_value = random.uniform(-max_val, -min_val)
                min_value = random.uniform(max_value, 0)
                most_likely = random.uniform(max_value, min_value)
            else:
                min_value = random.uniform(min_val, max_val)
                max_value = random.uniform(min_value, max_val * 2)
                most_likely = random.uniform(min_value, max_value)
            
            if not is_cost:
                min_value = math.ceil(min_value)
                most_likely = math.ceil(most_likely)
                max_value = math.ceil(max_value)
            
            return min_value, most_likely, max_value

        # Generate cost and delay values
        cost_min, cost_most_likely, cost_max = generate_pert_values(is_opportunity, 1000, 10000)
        delay_min, delay_most_likely, delay_max = generate_pert_values(is_opportunity, 1, 20, is_cost=False)

        # Create a risk dictionary
        risk = {
            "Risk ID": f"R{i:03d}",
            "Description": "",  # Placeholder for description
            "Initial Probability": round(random.uniform(0.1, 0.49), 2),  # Ensure probability is less than 0.5
            "Occurrence Time (Day)": random.randint(1, project_duration_days),
            "Cost PERT (min)": f"{cost_min:.1f}".replace(".", ","),
            "Cost PERT (most likely)": f"{cost_most_likely:.1f}".replace(".", ","),
            "Cost PERT (max)": f"{cost_max:.1f}".replace(".", ","),
            "Delay PERT (min)": delay_min,
            "Delay PERT (most likely)": delay_most_likely,
            "Delay PERT (max)": delay_max,
            "Dependent Risks": "",
            "Probability Adjustment": "",
            "Is Opportunity": is_opportunity
        }
        risks.append(risk)
    
    # Sort risks by occurrence time
    risks.sort(key=lambda x: x["Occurrence Time (Day)"])
    
    # Now assign descriptions after sorting
    for i, risk in enumerate(risks):
        risk["Description"] = f"{'Opportunity' if risk['Is Opportunity'] else 'Threat'} {i + 1} - Description"
        risk["Risk ID"] = f"R{i + 1:03d}"  # Ensure Risk ID matches the order

    # Generate dependencies and probability adjustments
    for i, risk in enumerate(risks):
        if i < len(risks) - 1:  # Skip the last risk as it can't have dependencies
            potential_future_risks = risks[i+1:]
            num_dependencies = random.randint(0, min(3, len(potential_future_risks)))
            if num_dependencies > 0:
                dependencies = random.sample(potential_future_risks, num_dependencies)
                risk["Dependent Risks"] = ", ".join([dep["Risk ID"] for dep in dependencies])
                
                # Set probability adjustment based on whether the risk is an opportunity or threat
                adjustment = round(random.uniform(0.05, 0.2), 2)
                if risk["Is Opportunity"]:
                    adjustment = -adjustment  # Decrease probability for opportunities
                risk["Probability Adjustment"] = adjustment
            else:
                risk["Probability Adjustment"] = ""
    
    # Remove the temporary "Is Opportunity" key
    for risk in risks:
        del risk["Is Opportunity"]
    
    return risks

def create_excel_file(risks, filename="risk_data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Data"

    headers = [
        "Risk ID", "Description", "Initial Probability", "Occurrence Time (Day)", 
        "Cost PERT (min)", "Cost PERT (most likely)", "Cost PERT (max)",
        "Delay PERT (min)", "Delay PERT (most likely)", "Delay PERT (max)",
        "Dependent Risks", "Probability Adjustment"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row, risk in enumerate(risks, start=2):
        for col, (key, value) in enumerate(risk.items(), start=1):
            ws.cell(row=row, column=col, value=value)

    # Add notes about units and PERT terminology
    ws['D1'].comment = openpyxl.comments.Comment('Days since project start', 'Auto-generated')
    ws['E1'].comment = openpyxl.comments.Comment('Cost values in monetary units. min: Minimum, most likely: Most likely, max: Maximum', 'Auto-generated')
    ws['H1'].comment = openpyxl.comments.Comment('Delay values in whole days. min: Minimum, most likely: Most likely, max: Maximum', 'Auto-generated')

    wb.save(filename)
    print(f"Excel file '{filename}' has been created successfully.")

if __name__ == "__main__":
    num_risks = 15  # You can change this number to generate more or fewer risks
    project_duration_days = 365  # Assuming a one-year project, adjust as needed
    risks = generate_risk_data(num_risks, project_duration_days)
    create_excel_file(risks)